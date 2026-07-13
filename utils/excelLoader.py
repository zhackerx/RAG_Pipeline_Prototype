from pathlib import Path

from openpyxl import load_workbook


class ExcelLoader:
    def __init__(self, file_path: str | Path):
        self.file_path = Path(file_path)

    def load_as_text(self) -> str:
        workbook = load_workbook(filename=str(self.file_path), data_only=True)
        lines: list[str] = []

        for sheet in workbook.worksheets:
            lines.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if cells:
                    lines.append(" | ".join(cells))

        return "\n".join(lines)
